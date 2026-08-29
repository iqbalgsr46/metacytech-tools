import os
import json
from datetime import datetime
from typing import Dict, Any, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .base_reporter import BaseReporter
from ..config import RESULTS_DIR

class ExcelReporter(BaseReporter):
    """Export OSINT scan results to professional Excel (.xlsx) with 12+ sheets"""
    
    def __init__(self):
        super().__init__()
        self.name = "ExcelReporter"
        
    def export(self, data: Dict[str, Any], filepath: str = None) -> str:
        if not HAS_OPENPYXL:
            return "ERROR: openpyxl belum terinstall. Jalankan: pip install openpyxl"
        
        if not filepath:
            target_type = data.get("target_type", "unknown")
            target_val = data.get("target", "unknown")
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = os.path.join(RESULTS_DIR, f"osint_{target_val}_{ts}.xlsx")
        
        try:
            wb = Workbook()
            self._create_dashboard(wb, data)
            self._create_sheet(wb, "DATA_KEPENDUDUKAN", data.get("kependudukan", {}))
            self._create_sheet(wb, "KELUARGA", data.get("keluarga", {}))
            self._create_sheet(wb, "DOKUMEN_BOCOR", data.get("dokumen_bocor", {}))
            self._create_sheet(wb, "BREACH_DATABASE", data.get("breach_db", {}))
            self._create_sheet(wb, "JEJAK_DIGITAL", data.get("jejak_digital", {}))
            self._create_sheet(wb, "MEDIA_SOSIAL", data.get("media_sosial", {}))
            self._create_sheet(wb, "FINANSIAL", data.get("finansial", {}))
            self._create_sheet(wb, "KOMUNIKASI", data.get("komunikasi", {}))
            self._create_sheet(wb, "PENDIDIKAN", data.get("pendidikan", {}))
            self._create_sheet(wb, "DARK_WEB", data.get("dark_web", {}))
            self._create_sheet(wb, "SUMBER", data.get("sumber", {}))
            
            wb.save(filepath)
            return f"OK: {filepath}"
            
        except Exception as e:
            return f"ERROR: {e}"
    
    def _create_dashboard(self, wb: Workbook, data: Dict):
        ws = wb.active
        ws.title = "DASHBOARD"
        
        # Header styling
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        ws.cell(row=1, column=1, value="OSINT ILEGAL — SCAN REPORT").font = Font(bold=True, size=16, color="002060")
        ws.merge_cells("A1:D1")
        
        ws.cell(row=3, column=1, value="Timestamp").font = Font(bold=True)
        ws.cell(row=3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        ws.cell(row=4, column=1, value="Target").font = Font(bold=True)
        ws.cell(row=4, column=2, value=data.get("target", "N/A"))
        
        ws.cell(row=5, column=1, value="Target Type").font = Font(bold=True)
        ws.cell(row=5, column=2, value=data.get("target_type", "N/A"))
        
        # Stats
        ws.cell(row=7, column=1, value="SCAN STATISTICS").font = Font(bold=True, size=13, color="002060")
        headers = ["Category", "Data Points", "Status", "Confidence"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=8, column=i, value=h)
            cell.fill = header_fill
            cell.font = header_font
        
        all_data = data.get("all_results", data)
        row = 9
        for category, items in all_data.items():
            if isinstance(items, dict):
                ws.cell(row=row, column=1, value=category)
                ws.cell(row=row, column=2, value=len(items))
                ws.cell(row=row, column=3, value="FOUND" if items else "NONE")
                ws.cell(row=row, column=4, value="LOW" if items else "N/A")
                row += 1
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
    
    def _create_sheet(self, wb: Workbook, title: str, data: Dict):
        if not data:
            ws = wb.create_sheet(title=title)
            ws.cell(row=1, column=1, value=f"No {title} data found")
            return
        
        ws = wb.create_sheet(title=title)
        header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        
        if isinstance(data, dict) and "fields" in data:
            fields = data.get("fields", {})
            ws.cell(row=1, column=1, value="Field").font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=2, value="Value").font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            ws.cell(row=1, column=3, value="Confidence").font = header_font
            ws.cell(row=1, column=3).fill = header_fill
            
            for i, (k, v) in enumerate(fields.items(), 2):
                ws.cell(row=i, column=1, value=k.replace("_", " ").title())
                ws.cell(row=i, column=2, value=str(v)[:100])
                ws.cell(row=i, column=3, value=data.get("confidence", "LOW"))
                if i % 2 == 0:
                    for c in range(1, 4):
                        ws.cell(row=i, column=c).fill = alt_fill
        elif isinstance(data, list):
            for i, item in enumerate(data, 1):
                if isinstance(item, dict):
                    for j, (k, v) in enumerate(item.items(), 1):
                        ws.cell(row=i, column=j, value=v)
        
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 15
        ws.auto_filter.ref = ws.dimensions if ws.max_row > 1 else None
        ws.freeze_panes = "A2"
